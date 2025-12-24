from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from typing import Any, Dict, List
from datetime import datetime
from ...base import BaseExportService

class XLSXListExportService(BaseExportService):
    """Servicio para exportar una lista de reportes en un archivo XLSX"""

    def __init__(self):
        self.workbook = None
        self.colors = {
            "header_fill": "D9D9D9",
            "row_odd_fill": "FFFFFF",
            "row_even_fill": "F2F2F2",
        }
        self.border = Border(
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF")
        )

    def generate_file(self, data: Any, options: Dict = None) -> io.BytesIO:
        if not self.validate_data(data):
            raise ValueError("Datos no válidos")

        if not isinstance(data, list):
            raise ValueError("Para listado se espera una lista de reportes")

        self.workbook = Workbook()
        sheet = self.workbook.active
        sheet.title = "Listado Reportes"

        # Encabezados conforme al modelo completo
        headers = [
            ("id", "ID de reporte"),
            ("user.documentId", "Documento de usuario"),
            ("user.userInformation.name", "Nombre del usuario"),
            ("user.userInformation.lastName", "Apellido del usuario"),
            ("externalNameUser", "Nombre externo del usuario"),
            ("externalOrganization", "Organizacion externa"),
            ("reportTitle", "Titulo del reporte"),
            ("conversation", "Conversacion"),
            ("base", "Base"),
            ("createdAt", "Fecha de creacion"),
            ("updatedAt", "Fecha de actualizacion"),
            ("unity", "Unidad"),
            ("rig", "Equipo (rig)"),
            ("project", "Proyecto"),
            ("field", "Campo"),
            ("reportType", "Tipo de reporte"),
            ("hazardClassification", "Clasificacion del peligro"),
            ("hazardType", "Tipo de peligro"),
            ("detailedDescription", "Descripcion detallada"),
            ("findingCause", "Causa del hallazgo"),
            ("reportEvidence", "Evidencias del reporte"),
            ("actions", "Acciones"),
            ("reportStatus", "Estado del reporte"),
            ("loraReportCode", "Codigo de reporte LORA"),
        ]

        # Escribir encabezado
        for col_index, (_, header_label) in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header_label)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors["header_fill"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border
        sheet.row_dimensions[1].height = 20

        # Escribir filas
        for row_index, report in enumerate(data, start=2):
            fill_color = self.colors["row_even_fill"] if row_index % 2 == 0 else self.colors["row_odd_fill"]
            for col_index, (header_key, _) in enumerate(headers, start=1):
                # permitir navegación en campos anidados (e.g., user.externalName)
                value = self._format_value(self._get_nested_value(report, header_key))
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.fill = PatternFill(start_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = self.border
            sheet.row_dimensions[row_index].height = 30

        # Ajustar ancho de columnas
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    length = len(str(cell.value or ""))
                    if length > max_length:
                        max_length = length
                except:
                    pass
            sheet.column_dimensions[column].width = min(max_length + 5, 50)

        # Crear buffer
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def _get_nested_value(self, obj: Dict, key: str) -> Any:
        """
        Permite acceder a valores anidados usando punto como separador,
        por ejemplo: 'user.externalName'
        """
        parts = key.split('.')
        value = obj
        for part in parts:
            if isinstance(value, dict):
                value = value.get(part, None)
            else:
                value = None
            if value is None:
                return ""
        return value

    def _format_value(self, value: Any) -> Any:
        """
        Normaliza valores complejos (listas/diccionarios) para que se vean legibles en la celda.
        """
        if isinstance(value, list):
            formatted_items = []
            for item in value:
                if isinstance(item, dict):
                    parts = [f"{k}: {v}" for k, v in item.items() if v not in (None, "")]
                    formatted_items.append(", ".join(parts) if parts else str(item))
                else:
                    formatted_items.append(str(item))
            return "\n".join(formatted_items)
        if isinstance(value, dict):
            return ", ".join([f"{k}: {v}" for k, v in value.items() if v not in (None, "")])
        return value

    def get_content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def get_file_extension(self) -> str:
        return ".xlsx"
