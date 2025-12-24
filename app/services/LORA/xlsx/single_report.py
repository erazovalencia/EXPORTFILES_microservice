from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from typing import Any, Dict, List
from datetime import datetime
from ...base import BaseExportService


class XLSXExportService(BaseExportService):
    """Servicio para generar reportes XLSX con formato empresarial estructurado"""

    FIELDS = [
        ("id", "ID de reporte"),
        ("userId", "ID de usuario"),
        ("user.documentId", "Documento de usuario"),
        ("user.userInformation.name", "Nombre del usuario"),
        ("user.userInformation.lastName", "Apellido del usuario"),
        ("externalNameUser", "Nombre externo del usuario"),
        ("externalOrganization", "Organizacion externa"),
        ("reportTitle", "Titulo del reporte"),
        ("conversation", "Conversacion"),
        ("base", "Base"),
        ("createdAt", "Fecha de creacion"),
        ("updatedAt", "Fecha de actualizacion"),
        ("unity", "Unidad"),
        ("rig", "Equipo (rig)"),
        ("project", "Proyecto"),
        ("field", "Campo"),
        ("reportType", "Tipo de reporte"),
        ("hazardClassification", "Clasificacion del peligro"),
        ("hazardType", "Tipo de peligro"),
        ("detailedDescription", "Descripcion detallada"),
        ("findingCause", "Causa del hallazgo"),
        ("reportEvidence", "Evidencias del reporte"),
        ("actions", "Acciones"),
        ("reportStatus", "Estado del reporte"),
        ("loraReportCode", "Codigo de reporte LORA"),
    ]

    def __init__(self):
        self.workbook = None
        # Colores y estilos
        self.colors = {
            "header_fill": "DCE6F1",
            "summary_fill": "F2F2F2",
            "open_fill": "F8CECC",   # rojo claro
            "close_fill": "D5E8D4",  # verde claro
        }
        self.border = Border(
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF")
        )

    def generate_file(self, data: Any, options: Dict = None) -> io.BytesIO:
        """Genera un archivo XLSX con diseño estructurado"""
        if not self.validate_data(data):
            raise ValueError("Los datos proporcionados no son válidos")

        self.workbook = Workbook()
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

        if isinstance(data, dict):
            self._create_summary_sheet(data)
            self._create_details_sheet(data)
            self._create_actions_sheet(data)
        else:
            self._create_simple_sheet(data)

        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer

    # ---------------------------
    # HOJA 1: RESUMEN
    # ---------------------------
    def _create_summary_sheet(self, data: Dict):
        ws = self.workbook.create_sheet("Resumen")
        ws.sheet_properties.tabColor = "1F497D"

        ws.merge_cells("A1:B1")
        title = ws["A1"]
        title.value = data.get("reportTitle", "Reporte de Hallazgo").upper()
        title.font = Font(bold=True, size=14, color="1F497D")
        title.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 25

        # Campos clave
        summary_fields = [
            (label, self._format_value(self._get_nested_value(data, key, fallback=data)))
            for key, label in self.FIELDS
        ]

        start_row = 3
        ws.append(["Campo", "Valor"])
        ws["A3"].font = ws["B3"].font = Font(bold=True)
        ws["A3"].fill = ws["B3"].fill = PatternFill(start_color=self.colors["header_fill"], fill_type="solid")
        ws["A3"].alignment = ws["B3"].alignment = Alignment(horizontal="center")

        for i, (label, value) in enumerate(summary_fields, start=start_row + 1):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = str(value)
            ws[f"A{i}"].fill = PatternFill(start_color=self.colors["summary_fill"], fill_type="solid")
            ws[f"A{i}"].font = Font(bold=True)
            ws[f"A{i}"].border = ws[f"B{i}"].border = self.border

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 60

    # ---------------------------
    # HOJA 2: DETALLES
    # ---------------------------
    def _create_details_sheet(self, data: Dict):
        ws = self.workbook.create_sheet("Detalles")
        ws.sheet_properties.tabColor = "76933C"

        sections = [
            ("Descripción", data.get("detailedDescription")),
            ("Causa", data.get("findingCause")),
            ("Conversación", data.get("conversation")),
            ("Evidencias", "\n".join(data.get("evidence", [])) if data.get("evidence") else "Sin evidencias adjuntas"),
        ]

        row = 1
        for title, content in sections:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            cell = ws.cell(row=row, column=1, value=title.upper())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4BACC6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 22

            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=3)
            text_cell = ws.cell(row=row, column=1, value=content or "N/A")
            text_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="justify")
            text_cell.border = self.border
            ws.row_dimensions[row].height = 80
            row += 5

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 45

    # ---------------------------
    # HOJA 3: ACCIONES
    # ---------------------------
    def _create_actions_sheet(self, data: Dict):
        ws = self.workbook.create_sheet("Acciones")
        ws.sheet_properties.tabColor = "C0504D"

        headers = ["Descripción", "Responsable", "Fecha Límite", "Estado"]
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="404040", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        actions = data.get("actions", [])
        if not actions:
            ws.append(["Sin acciones registradas"])
        else:
            for i, act in enumerate(actions, start=2):
                status = act.get("status", "open").lower()
                fill_color = self.colors["close_fill"] if status == "close" else self.colors["open_fill"]

                ws.cell(row=i, column=1, value=act.get("description", "N/A"))
                ws.cell(row=i, column=2, value=act.get("responsible", "No asignado"))
                ws.cell(row=i, column=3, value=act.get("dueDate", "N/A"))
                ws.cell(row=i, column=4, value=status.upper())

                for col in range(1, 5):
                    c = ws.cell(row=i, column=col)
                    c.alignment = Alignment(vertical="center", wrap_text=True)
                    c.border = self.border
                    c.fill = PatternFill(start_color=fill_color, fill_type="solid")

        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15

    # ---------------------------
    # HOJA SIMPLE (fallback)
    # ---------------------------
    def _create_simple_sheet(self, data: Any):
        ws = self.workbook.create_sheet("Datos")
        ws["A1"] = str(data)
        ws["A1"].font = Font(bold=True)
        ws.column_dimensions["A"].width = 50

    # ---------------------------
    # HELPERS
    # ---------------------------
    def _get_nested_value(self, obj: Dict, key: str, fallback: Dict = None) -> Any:
        parts = key.split('.')
        value = obj
        for part in parts:
            if isinstance(value, dict):
                value = value.get(part, None)
            else:
                value = None
            if value is None:
                break
        if value is None and fallback and key == "reportEvidence":
            evidences = fallback.get("evidence", [])
            return evidences if evidences else None
        if value is None and fallback and key == "actions":
            return fallback.get("actions", [])
        return value

    def _format_value(self, value: Any) -> Any:
        if value is None:
            return "N/A"
        if isinstance(value, list):
            formatted_items = []
            for item in value:
                if isinstance(item, dict):
                    parts = [f"{k}: {v}" for k, v in item.items() if v not in (None, "")]
                    formatted_items.append(", ".join(parts) if parts else str(item))
                else:
                    formatted_items.append(str(item))
            return "\n".join(formatted_items) if formatted_items else "N/A"
        if isinstance(value, dict):
            parts = [f"{k}: {v}" for k, v in value.items() if v not in (None, "")]
            return ", ".join(parts) if parts else "N/A"
        return value

    # ---------------------------
    # METADATOS
    # ---------------------------
    def get_content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def get_file_extension(self) -> str:
        return ".xlsx"
