from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from typing import Any, Dict, List
from ..services.base import BaseExportService


class XLSXExportService(BaseExportService):
    """Servicio para generar archivos Excel XLSX"""
    
    def __init__(self):
        self.workbook = None
    
    def generate_file(self, data: Any, options: Dict = None) -> io.BytesIO:
        """Genera un archivo XLSX"""
        if not self.validate_data(data):
            raise ValueError("Los datos proporcionados no son válidos")
        
        self.workbook = Workbook()
        
        # Eliminar la hoja por defecto
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)
        
        # Si los datos son un diccionario con estructura de documento
        if isinstance(data, dict):
            self._process_document_data(data, options or {})
        else:
            # Crear hoja simple con los datos
            self._create_simple_sheet(data)
        
        # Guardar en buffer
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _process_document_data(self, data: Dict, options: Dict):
        """Procesa datos estructurados de documento"""
        # Si hay tablas múltiples, crear una hoja por tabla
        if 'tables' in data and isinstance(data['tables'], list):
            for i, table in enumerate(data['tables']):
                sheet_name = table.get('title', f'Tabla_{i+1}')
                self._create_table_sheet(table, sheet_name)
        # Si es una tabla simple
        elif 'headers' in data and 'rows' in data:
            sheet_name = data.get('title', 'Datos')
            self._create_table_sheet(data, sheet_name)
        else:
            # Crear hoja con información general
            self._create_info_sheet(data)
    
    def _create_table_sheet(self, table_data: Dict, sheet_name: str):
        """Crea una hoja con una tabla"""
        headers = table_data.get('headers', [])
        rows = table_data.get('rows', [])
        
        if not headers or not rows:
            return
        
        # Crear nueva hoja
        worksheet = self.workbook.create_sheet(title=sheet_name[:31])  # Excel limita a 31 caracteres
        
        # Añadir encabezados
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Añadir filas de datos
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, cell_data in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=cell_data)
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_info_sheet(self, data: Dict):
        """Crea una hoja con información general"""
        worksheet = self.workbook.create_sheet(title="Información")
        
        row = 1
        if 'title' in data and data['title']:
            cell = worksheet.cell(row=row, column=1, value=data['title'])
            cell.font = Font(bold=True, size=14)
            row += 2
        
        if 'content' in data and isinstance(data['content'], list):
            for content in data['content']:
                worksheet.cell(row=row, column=1, value=content)
                row += 1
        
        # Ajustar ancho de la primera columna
        worksheet.column_dimensions['A'].width = 50
    
    def _create_simple_sheet(self, data: Any):
        """Crea una hoja simple con datos"""
        worksheet = self.workbook.create_sheet(title="Datos")
        
        if isinstance(data, (list, tuple)):
            for i, item in enumerate(data, 1):
                worksheet.cell(row=i, column=1, value=str(item))
        else:
            worksheet.cell(row=1, column=1, value=str(data))
        
        worksheet.column_dimensions['A'].width = 30
    
    def get_content_type(self) -> str:
        """Retorna el tipo de contenido MIME para XLSX"""
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    def get_file_extension(self) -> str:
        """Retorna la extensión del archivo XLSX"""
        return ".xlsx"